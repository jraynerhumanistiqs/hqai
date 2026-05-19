"""Build the HQ.ai pilot feedback tracker as a Google-Sheet-ready .xlsx.

Run: py -3 scripts/build-pilot-feedback-sheet.py

Output: docs/HQ-AI-PILOT-FEEDBACK.xlsx

Three sheets:
  - How to use
  - Feedback Log     (main tracker, autofilter + frozen header)
  - Surface Scores   (per-surface 1-5 + the five lenses)

Designed to be uploaded to Google Drive (right click in Drive -> Open
with Google Sheets). All data validation + freeze panes + autofilters
transfer cleanly to Google Sheets format.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date
from pathlib import Path

# ── Brand palette ────────────────────────────────────────────────
ACCENT     = 'FFD97757'  # Clay
ACCENT_DK  = 'FFC6613F'
INK        = 'FF141413'
INK_SOFT   = 'FF3D3D3A'
INK_MUTED  = 'FF5E5D59'
CREAM      = 'FFFAF9F5'
CREAM_SOFT = 'FFF0EEE6'
WHITE      = 'FFFFFFFF'
BORDER     = 'FFD1CFC5'

FONT_BODY    = Font(name='Arial', size=10, color=INK)
FONT_BODY_S  = Font(name='Arial', size=9,  color=INK_MUTED)
FONT_BOLD    = Font(name='Arial', size=10, color=INK,  bold=True)
FONT_HEADER  = Font(name='Arial', size=10, color=WHITE, bold=True)
FONT_TITLE   = Font(name='Arial', size=18, color=INK,  bold=True)
FONT_H2      = Font(name='Arial', size=13, color=INK,  bold=True)

FILL_HEADER  = PatternFill('solid', start_color=ACCENT,    end_color=ACCENT)
FILL_HEADER_DK = PatternFill('solid', start_color=ACCENT_DK, end_color=ACCENT_DK)
FILL_CREAM   = PatternFill('solid', start_color=CREAM,     end_color=CREAM)
FILL_CREAM_SOFT = PatternFill('solid', start_color=CREAM_SOFT, end_color=CREAM_SOFT)
FILL_WHITE   = PatternFill('solid', start_color=WHITE,     end_color=WHITE)

ALIGN_WRAP   = Alignment(horizontal='left', vertical='top', wrap_text=True)
ALIGN_HEADER = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

THIN = Side(style='thin', color=BORDER)
BORDER_ALL = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

# Surfaces match the 11 cards in the testing guide (post CV-Formatter removal).
SURFACES = [
    'Sign in',
    'Onboarding wizard',
    'Dashboard home',
    'AI Advisor',
    'AI Administrator',
    'CV Scoring Agent',
    'Shortlist',
    'Campaign Coach',
    'My Documents',
    'Settings',
    '$25 Letter of Offer',
]
SEVERITY = ['Show-stopper', 'Annoying', 'Polish']
CATEGORY = ['Bug', 'UX / clarity', 'Copy / language', 'Performance', 'Feature gap', 'Design polish', 'AI quality']
STATUS   = ['New', 'Triaged', 'In progress', 'Fixed', "Won't fix", 'Duplicate']

LENSES = [
    'Would a real SME owner click this?',
    'Does it save you 80%?',
    'What is the one missing feature?',
    'Where do you mistrust the AI?',
    'What would your clients break?',
]

# ── Build ────────────────────────────────────────────────────────
wb = Workbook()
ws_intro = wb.active
ws_intro.title = 'How to use'
ws_log = wb.create_sheet('Feedback Log')
ws_scores = wb.create_sheet('Surface Scores')

# ──────────────────────────── SHEET 1: How to use ────────────────────
ws = ws_intro
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 110

def title_row(row, text, fill=FILL_HEADER, font=FONT_HEADER, height=28):
    cell = ws.cell(row=row, column=2, value=text)
    cell.fill = fill
    cell.font = font
    cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = height

def body_row(row, text, bold=False, italic=False, indent=False):
    cell = ws.cell(row=row, column=2, value=text)
    cell.font = Font(name='Arial', size=10, color=INK, bold=bold, italic=italic)
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=1 if indent else 0)
    cell.fill = FILL_CREAM
    ws.row_dimensions[row].height = 22 if not text else max(22, 18 * (len(text) // 110 + 1))

def spacer(row):
    ws.row_dimensions[row].height = 8

# Title block
ws.cell(row=2, column=2, value='HQ.ai Pilot Feedback').font = FONT_TITLE
ws.cell(row=2, column=2).alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws.row_dimensions[2].height = 34
ws.cell(row=3, column=2, value='How to use this sheet').font = Font(name='Arial', size=11, color=ACCENT_DK, bold=True, italic=True)
ws.cell(row=3, column=2).alignment = Alignment(horizontal='left', vertical='center', indent=1)

row = 5
title_row(row, 'The basics'); row += 1
for line in [
    '1.  One row per piece of feedback. Keep them small and specific - one issue per row, not a paragraph of mixed thoughts.',
    '2.  Fill in the columns from left to right as you find issues. The dropdowns make it fast.',
    '3.  Triage columns (Status and Owner) are for Jimmy to fill - leave them blank as the tester.',
    '4.  Pair this with the Pilot Testing Map PDF. Page 2 of the PDF shows side-by-side examples of useful vs not-useful feedback - worth a 30 second read before you start.',
    '5.  Stuck on whether something is worth logging? Log it. Better to surface and dismiss later than miss something real.',
]:
    body_row(row, line); row += 1

spacer(row); row += 1
title_row(row, 'What each column is for', fill=FILL_HEADER_DK); row += 1
for label, desc in [
    ('Surface',            'Which page you were on. Use the dropdown - matches the 11 surfaces in the Pilot Testing Map.'),
    ('Severity',           'Show-stopper / Annoying / Polish. See the legend below.'),
    ('Category',           'Bug / UX / Copy / Performance / Feature gap / Design polish / AI quality.'),
    ('What you did',       'The steps you took. Plain language - "I uploaded a CV and clicked Score" is enough.'),
    ('What happened',      'The actual behaviour you saw. Be specific - "the AI did not include the candidate name in the body" beats "the doc looked wrong".'),
    ('What you expected',  'Only fill this in if it is not obvious from What happened.'),
    ('Suggested fix',      'Optional. No design pressure - we do not expect you to fix it, just observe it.'),
    ('Screenshot or Loom', 'Paste a URL. Drive / Loom links are gold. Cannot attach a file directly to a Google Sheet cell.'),
    ('Browser / Device',   'Chrome desktop / Safari iPhone / Edge laptop - whatever you were using.'),
]:
    body_row(row, f'{label}  -  {desc}', bold=False); row += 1

spacer(row); row += 1
title_row(row, 'Severity legend'); row += 1
for label, desc in [
    ('Show-stopper', 'Stops you doing the task. Cannot sign in. Cannot generate a document. Hard error with no recovery.'),
    ('Annoying',     'Slows you down or confuses. Hover state wrong, dropdown closes too fast, copy reads weirdly, missing labels.'),
    ('Polish',       'Cosmetic. Spacing, alignment, colour, font weight. Nice-to-have. Log it last.'),
]:
    body_row(row, f'{label}  -  {desc}', bold=False); row += 1

spacer(row); row += 1
title_row(row, 'Category legend', fill=FILL_HEADER_DK); row += 1
for label, desc in [
    ('Bug',             'Something is broken. Error message, blank page, wrong output.'),
    ('UX / clarity',    'Confusing flow, hidden button, unclear next step, surprise modal.'),
    ('Copy / language', 'Wording reads wrong. Typos. HR jargon a small business owner would not know.'),
    ('Performance',     'Slow load, slow response, freezing.'),
    ('Feature gap',     'Something you expected to find that is not there.'),
    ('Design polish',   'Visual issues - spacing, alignment, colour, type weight.'),
    ('AI quality',      'The AI output is wrong, generic, biased, or missing critical detail.'),
]:
    body_row(row, f'{label}  -  {desc}', bold=False); row += 1

spacer(row); row += 1
title_row(row, 'Status legend (Jimmy fills)'); row += 1
for label, desc in [
    ('New',         'Just logged, not yet triaged.'),
    ('Triaged',     'Reviewed, going into the backlog.'),
    ('In progress', 'Being worked on now.'),
    ('Fixed',       'Done. Ship verified.'),
    ("Won't fix",   'Logged and decided against. Reason in the Notes / Owner column.'),
    ('Duplicate',   'Same issue already logged elsewhere - point to the duplicate row.'),
]:
    body_row(row, f'{label}  -  {desc}', bold=False); row += 1

spacer(row); row += 1
title_row(row, 'Reminder: the five questions we most want answered', fill=FILL_HEADER_DK); row += 1
for q in LENSES:
    body_row(row, q); row += 1
spacer(row); row += 1
body_row(row, 'Score them on the Surface Scores tab. One row per question - 1 to 5 plus a one-liner.', italic=True)

# ──────────────────────────── SHEET 2: Feedback Log ──────────────
ws = ws_log
ws.sheet_view.showGridLines = False
ws.freeze_panes = 'A2'

headers = [
    ('ID',                       8),
    ('Date logged',              12),
    ('Tester',                   14),
    ('Surface',                  22),
    ('Severity',                 14),
    ('Category',                 18),
    ('What you did',             40),
    ('What happened',            40),
    ('What you expected',        32),
    ('Suggested fix (optional)', 32),
    ('Screenshot or Loom link',  28),
    ('Browser / Device',         18),
    ('Status (Jimmy fills)',     16),
    ('Owner / Notes (Jimmy fills)', 22),
]
for col_idx, (label, width) in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=label)
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.alignment = ALIGN_HEADER
    cell.border = BORDER_ALL
    ws.column_dimensions[get_column_letter(col_idx)].width = width
ws.row_dimensions[1].height = 34

# Worked example row (row 2)
example = [
    1,
    date.today(),
    'Bianca',
    'AI Administrator',
    'Annoying',
    'AI quality',
    'Generated a Letter of Offer with the candidate name Sarah Chen',
    "The body of the letter says 'the successful candidate' instead of using her name",
    'Her name to appear in the body prose, not just the recipient block',
    '',
    '',
    'Chrome desktop',
    'New',
    '',
]
for col_idx, value in enumerate(example, start=1):
    cell = ws.cell(row=2, column=col_idx, value=value)
    cell.fill = FILL_CREAM_SOFT
    cell.font = Font(name='Arial', size=10, color=INK_SOFT, italic=True)
    cell.alignment = ALIGN_WRAP
    cell.border = BORDER_ALL
ws.cell(row=2, column=2).number_format = 'dd/mm/yyyy'
ws.row_dimensions[2].height = 48
# Note in the ID col cell tooltip
ws.cell(row=2, column=1).comment = None

# 50 blank rows pre-formatted ready for entry. Auto-fill ID via formula
# so testers do not have to count.
for r in range(3, 53):
    ws.cell(row=r, column=1, value=f'=ROW()-1')
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=r, column=col_idx)
        cell.alignment = ALIGN_WRAP
        cell.font = FONT_BODY
        cell.border = BORDER_ALL
        cell.fill = FILL_WHITE
    ws.cell(row=r, column=2).number_format = 'dd/mm/yyyy'
    ws.row_dimensions[r].height = 28

# Data validation dropdowns
def quoted(items):
    return '"' + ','.join(items) + '"'

dv_surface  = DataValidation(type='list', formula1=quoted(SURFACES), allow_blank=True)
dv_severity = DataValidation(type='list', formula1=quoted(SEVERITY), allow_blank=True)
dv_category = DataValidation(type='list', formula1=quoted(CATEGORY), allow_blank=True)
dv_status   = DataValidation(type='list', formula1=quoted(STATUS),   allow_blank=True)
for dv in [dv_surface, dv_severity, dv_category, dv_status]:
    dv.showDropDown = False  # show the arrow (Excel inverted flag)
    ws.add_data_validation(dv)
dv_surface.add('D2:D200')
dv_severity.add('E2:E200')
dv_category.add('F2:F200')
dv_status.add('M2:M200')

# Autofilter on header row covering data range
ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}52'

# ──────────────────────────── SHEET 3: Surface Scores ────────────
ws = ws_scores
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 26
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 56

# Title
ws.cell(row=2, column=2, value='Surface Scores').font = FONT_TITLE
ws.cell(row=2, column=2).alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[2].height = 30
sub = ws.cell(row=3, column=2, value='Rate each surface 1 (needs work) to 5 (client-ready), then answer the five questions.')
sub.font = Font(name='Arial', size=10, color=INK_MUTED, italic=True)
ws.merge_cells('B3:E3')

# Table A header
ROW_A = 5
hdr = [('Surface', 'B'), ('Tester', 'C'), ('Rating 1-5', 'D'), ('One-sentence why', 'E')]
for label, col in hdr:
    c = ws[f'{col}{ROW_A}']
    c.value = label
    c.font = FONT_HEADER
    c.fill = FILL_HEADER
    c.alignment = ALIGN_HEADER
    c.border = BORDER_ALL
ws.row_dimensions[ROW_A].height = 26

for i, surface in enumerate(SURFACES):
    r = ROW_A + 1 + i
    ws.cell(row=r, column=2, value=surface).font = FONT_BOLD
    for col_idx in range(2, 6):
        cell = ws.cell(row=r, column=col_idx)
        cell.alignment = ALIGN_WRAP
        cell.font = FONT_BODY if col_idx > 2 else FONT_BOLD
        cell.border = BORDER_ALL
        cell.fill = FILL_CREAM if col_idx == 2 else FILL_WHITE
    ws.row_dimensions[r].height = 26

# 1-5 dropdown
dv_rating = DataValidation(type='list', formula1='"1,2,3,4,5"', allow_blank=True)
ws.add_data_validation(dv_rating)
dv_rating.add(f'D{ROW_A + 1}:D{ROW_A + len(SURFACES)}')

# Table B - The five questions
ROW_B = ROW_A + len(SURFACES) + 3
title_b = ws.cell(row=ROW_B - 1, column=2, value='The five questions we most want answered')
title_b.font = FONT_H2
title_b.alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[ROW_B - 1].height = 26
ws.merge_cells(start_row=ROW_B - 1, start_column=2, end_row=ROW_B - 1, end_column=5)

hdr_b = [('Question', 'B'), ('Tester', 'C'), ('Score 1-5', 'D'), ('Answer', 'E')]
for label, col in hdr_b:
    c = ws[f'{col}{ROW_B}']
    c.value = label
    c.font = FONT_HEADER
    c.fill = FILL_HEADER_DK
    c.alignment = ALIGN_HEADER
    c.border = BORDER_ALL
ws.row_dimensions[ROW_B].height = 26

for i, q in enumerate(LENSES, start=1):
    r = ROW_B + i
    ws.cell(row=r, column=2, value=f'{i}.  {q}').font = FONT_BOLD
    for col_idx in range(2, 6):
        cell = ws.cell(row=r, column=col_idx)
        cell.alignment = ALIGN_WRAP
        cell.font = FONT_BODY if col_idx > 2 else FONT_BOLD
        cell.border = BORDER_ALL
        cell.fill = FILL_CREAM if col_idx == 2 else FILL_WHITE
    ws.row_dimensions[r].height = 40

dv_rating_b = DataValidation(type='list', formula1='"1,2,3,4,5"', allow_blank=True)
ws.add_data_validation(dv_rating_b)
dv_rating_b.add(f'D{ROW_B + 1}:D{ROW_B + len(LENSES)}')

# ── Save ─────────────────────────────────────────────────────────
out_path = Path(__file__).resolve().parents[1] / 'docs' / 'HQ-AI-PILOT-FEEDBACK.xlsx'
out_path.parent.mkdir(parents=True, exist_ok=True)
wb.save(out_path)
print(f'Wrote {out_path} ({out_path.stat().st_size} bytes)')
